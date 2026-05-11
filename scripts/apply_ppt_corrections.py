"""Apply the PPT-driven corrections to communications.json and the source Excel.

The PPT delivered on 11 May 2026 corrected sender/receiver/locations for ~30
existing letters and added the London – New York – Istanbul axis (Dosya
134-140). This script patches:

  * data/communications.json   (the diplomacy map data)
  * "Pace of Communication - PURE.xlsx"  (the canonical source spreadsheet)

For each Dosya:
  * if an entry already exists for that ISO date, the matching one is updated
    (preferring the one currently linked to the same Dosya number);
  * otherwise a new entry is appended at the chronologically correct spot.

After this script runs, ``scripts/build_documents.py`` re-applies its
``DOSYA_METADATA_OVERRIDES`` to lock the per-folder display values regardless
of date collisions.

Run from the repo root:
    .venv/bin/python scripts/apply_ppt_corrections.py
"""
from __future__ import annotations

import json
from pathlib import Path

import openpyxl


ROOT = Path(__file__).resolve().parent.parent
COMMS_PATH = ROOT / "data" / "communications.json"
EXCEL_PATH = ROOT / "Pace of Communication - PURE.xlsx"


# ---------------------------------------------------------------------------
# Source of truth for the PPT corrections.
#
# Each entry maps a Dosya number to:
#   date              – ISO date the letter should carry on the website
#   sender            – Full official name of the writer
#   receiver          – Full official name of the addressee
#   sender_location   – City the letter is *sent from*
#   receiver_location – City the letter is *received at*
#   type              – Optional; defaults to "letter"
#
# Dosya numbers not in this table keep their existing communication mapping.
# ---------------------------------------------------------------------------
CORRECTIONS: dict[int, dict] = {
    # Dosya 4 — filename says 21.02.1859; old override forced 1858. Restore.
    4: {
        "date": "1859-02-21",
        "sender": "Cardinal Pietro Gianelli",
        "receiver": "Ottoman Consul of Naples Sigmund Spitzer",
        "sender_location": "Sardinia",
        "receiver_location": "Naples",
    },
    12: {
        "date": "1858-03-17",
        "sender": "Ottoman Foreign Minister Keçecizade Fuad Pasha",
        "receiver": "Ottoman Consul of Turin Rüstem Mariani Bey",
        "sender_location": "Istanbul",
        "receiver_location": "Turin",
    },
    13: {
        "date": "1858-03-18",
        "sender": "Count of Salmour Rugiero Gabaleone",
        "receiver": "Ottoman Consul of Turin Rüstem Mariani Bey",
        "sender_location": "Turin",
        "receiver_location": "Turin",
    },
    16: {
        "date": "1858-03-24",
        "sender": "Ottoman Foreign Minister Keçecizade Fuad Pasha",
        "receiver": "Ottoman Consul of Turin Rüstem Mariani Bey",
        "sender_location": "Istanbul",
        "receiver_location": "Turin",
    },
    19: {
        "date": "1858-03-31",
        "sender": "Ottoman Foreign Minister Keçecizade Fuad Pasha",
        "receiver": "Ottoman Consul of Turin Rüstem Mariani Bey",
        "sender_location": "Istanbul",
        "receiver_location": "Turin",
    },
    21: {
        "date": "1858-04-14",
        "sender": "Ottoman Foreign Minister Mahmud Nedim Pasha",
        "receiver": "Ottoman Consul of Athens Halil Bey (Halil Şerif Pasha)",
        "sender_location": "Istanbul",
        "receiver_location": "Athens",
    },
    22: {
        "date": "1858-04-14",
        "sender": "Ottoman Foreign Minister Keçecizade Fuad Pasha",
        "receiver": "Ottoman Consul of Turin Rüstem Mariani Bey",
        "sender_location": "Istanbul",
        "receiver_location": "Turin",
    },
    25: {
        "date": "1861-07-04",
        "sender": "Ottoman Consulate of Venice Breganze",
        "receiver": "Ottoman Foreign Minister Mahmud Nedim Pasha",
        "sender_location": "Venice",
        "receiver_location": "Istanbul",
    },
    # Dosya 28 — filename says 01.12.1858; old override forced 1858-11-01.
    28: {
        "date": "1858-12-01",
        "sender": "Ottoman Foreign Minister Keçecizade Fuad Pasha",
        "receiver": "Ottoman Consul of Turin Rüstem Mariani Bey",
        "sender_location": "Istanbul",
        "receiver_location": "Turin",
    },
    35: {
        "date": "1859-02-09",
        "sender": "Ottoman Foreign Minister Keçecizade Fuad Pasha",
        "receiver": "Ottoman Consul of Naples Sigmund Spitzer",
        "sender_location": "Istanbul",
        "receiver_location": "Naples",
    },
    36: {
        "date": "1859-02-09",
        "sender": "Ottoman Consul of Turin Rüstem Mariani Bey",
        "receiver": "Ottoman Foreign Minister Keçecizade Fuad Pasha",
        "sender_location": "Turin",
        "receiver_location": "Istanbul",
    },
    37: {
        "date": "1859-02-15",
        "sender": "Cardinal Antonelli",
        "receiver": "Ottoman Consul of Turin Rüstem Mariani Bey",
        "sender_location": "Rome",
        "receiver_location": "Turin",
    },
    38: {
        "date": "1859-02-21",
        "sender": "Ottoman Consul of Naples Sigmund Spitzer",
        "receiver": "Ottoman Foreign Minister Keçecizade Fuad Pasha",
        "sender_location": "Naples",
        "receiver_location": "Istanbul",
    },
    39: {
        "date": "1859-02-21",
        "sender": "Legation of Bolognese Police (Cardinal Milesi)",
        "receiver": "Apostolic Nunciature of Turin",
        "sender_location": "Bologna",
        "receiver_location": "Turin",
    },
    40: {
        "date": "1859-02-23",
        "sender": "Apostolic Nunciature of Turin",
        "receiver": "Ottoman Consul of Turin Rüstem Mariani Bey",
        "sender_location": "Turin",
        "receiver_location": "Turin",
    },
    45: {
        "date": "1859-03-02",
        "sender": "Ottoman Consul of Turin Rüstem Mariani Bey",
        "receiver": "Ottoman Foreign Minister Keçecizade Fuad Pasha",
        "sender_location": "Turin",
        "receiver_location": "Istanbul",
    },
    48: {
        "date": "1859-03-10",
        "sender": "Ottoman Consul of Turin Rüstem Mariani Bey",
        "receiver": "Ottoman Foreign Minister Keçecizade Fuad Pasha",
        "sender_location": "Turin",
        "receiver_location": "Istanbul",
    },
    52: {
        "date": "1859-03-17",
        "sender": "Ottoman Consul of Turin Rüstem Mariani Bey",
        "receiver": "Ottoman Foreign Minister Keçecizade Fuad Pasha",
        "sender_location": "Turin",
        "receiver_location": "Istanbul",
    },
    53: {
        "date": "1859-03-23",
        "sender": "Ottoman Foreign Minister Keçecizade Fuad Pasha",
        "receiver": "Ottoman Consul of Naples Sigmund Spitzer",
        "sender_location": "Istanbul",
        "receiver_location": "Naples",
    },
    54: {
        "date": "1859-03-23",
        "sender": "Ottoman Foreign Minister Keçecizade Fuad Pasha",
        "receiver": "Ottoman Police Officer İkiades Bey",
        "sender_location": "Istanbul",
        "receiver_location": "Bologna",
    },
    55: {
        "date": "1859-03-23",
        "sender": "Ottoman Foreign Minister Keçecizade Fuad Pasha",
        "receiver": "Ottoman Consulate of Venice",
        "sender_location": "Istanbul",
        "receiver_location": "Venice",
    },
    60: {
        "date": "1859-03-31",
        "sender": "Ottoman Consul of Turin Rüstem Mariani Bey",
        "receiver": "Ottoman Foreign Minister Keçecizade Fuad Pasha",
        "sender_location": "Turin",
        "receiver_location": "Istanbul",
    },
    65: {
        "date": "1859-04-14",
        "sender": "Ottoman Consul of Turin Rüstem Mariani Bey",
        "receiver": "Ottoman Foreign Minister Keçecizade Fuad Pasha",
        "sender_location": "Turin",
        "receiver_location": "Istanbul",
    },
    68: {
        "date": "1859-04-27",
        "sender": "Ottoman Foreign Minister Keçecizade Fuad Pasha",
        "receiver": "Ottoman Consul of Turin Rüstem Mariani Bey",
        "sender_location": "Istanbul",
        "receiver_location": "Istanbul",
    },
    73: {
        "date": "1859-07-07",
        "sender": "Ottoman Consul of Turin Rüstem Mariani Bey",
        "receiver": "Ottoman Foreign Minister Keçecizade Fuad Pasha",
        "sender_location": "Turin",
        "receiver_location": "Istanbul",
    },
    80: {
        "date": "1859-11-02",
        "sender": "Ottoman Foreign Minister Keçecizade Fuad Pasha",
        "receiver": "Ottoman Consul of Naples Sigmund Spitzer",
        "sender_location": "Istanbul",
        "receiver_location": "Naples",
    },
    88: {
        "date": "1860-01-04",
        "sender": "Ottoman Foreign Minister Keçecizade Fuad Pasha",
        "receiver": "Ottoman Consul of Turin Rüstem Mariani Bey",
        "sender_location": "Istanbul",
        "receiver_location": "Istanbul",
    },
    95: {
        "date": "1860-04-19",
        "sender": "Ottoman Consul of Turin Rüstem Mariani Bey",
        "receiver": "Ottoman Foreign Minister Keçecizade Fuad Pasha",
        "sender_location": "Turin",
        "receiver_location": "Istanbul",
    },
    105: {
        "date": "1860-07-05",
        "sender": "Domenico Carutti, General Secretary of the Ministry of Foreign Affairs",
        "receiver": "Ottoman Consul of Turin Rüstem Mariani Bey",
        "sender_location": "Turin",
        "receiver_location": "Turin",
    },
    107: {
        "date": "1860-07-20",
        "sender": "Count of Cavour Camillo Benso",
        "receiver": "Ottoman Consul of Turin Rüstem Mariani Bey",
        "sender_location": "Turin",
        "receiver_location": "Turin",
    },
    111: {
        "date": "1860-08-23",
        "sender": "Ottoman Consul of Turin Rüstem Mariani Bey",
        "receiver": "Provisional Ottoman Foreign Minister Mehmet Esad Safvet Efendi",
        "sender_location": "Turin",
        "receiver_location": "Istanbul",
    },
    118: {
        "date": "1861-02-06",
        "sender": "Ottoman Foreign Minister Ali Pasha",
        "receiver": "Ottoman Consul of Turin Rüstem Mariani Bey",
        "sender_location": "Istanbul",
        "receiver_location": "Turin",
    },
    128: {
        "date": "1861-07-04",
        "sender": "Ottoman Consul of Turin Rüstem Mariani Bey",
        "receiver": "Ottoman Foreign Minister Ali Pasha",
        "sender_location": "Turin",
        "receiver_location": "Istanbul",
    },
    132: {
        "date": "1861-03-17",
        "sender": "Luigi Amedeo Melegari",
        "receiver": "Ottoman Consul of Turin Rüstem Mariani Bey",
        "sender_location": "Turin",
        "receiver_location": "Turin",
    },
    # ---- London – New York – Istanbul axis (Dosya 134-140) ----
    134: {
        "date": "1858-11-17",
        "sender": "Ottoman Foreign Minister Keçecizade Fuad Pasha",
        "receiver": "Ottoman Ambassador at London Kostaki Mousourus (Musurus) Bey",
        "sender_location": "Istanbul",
        "receiver_location": "London",
    },
    135: {
        "date": "1858-10-07",
        "sender": "Ottoman Consulate of New York, Hosford Smith",
        "receiver": "Ottoman Ambassador at London Kostaki Mousourus (Musurus) Bey",
        "sender_location": "New York",
        "receiver_location": "London",
    },
    136: {
        "date": "1858-10-08",
        "sender": "Ottoman Consulate of New York, Hosford Smith",
        "receiver": "Ottoman Foreign Minister Keçecizade Fuad Pasha",
        "sender_location": "New York",
        "receiver_location": "Istanbul",
    },
    137: {
        "date": "1858-11-17",
        "sender": "Ottoman Foreign Minister Keçecizade Fuad Pasha",
        "receiver": "Ottoman Consulate of New York, Hosford Smith",
        "sender_location": "Istanbul",
        "receiver_location": "New York",
    },
    138: {
        # Filename says 09.02.1859; PPT corrects to 07.02.1859.
        "date": "1859-02-07",
        "sender": "Ottoman Foreign Minister Keçecizade Fuad Pasha",
        "receiver": "Ottoman Ambassador at London Kostaki Mousourus (Musurus) Bey",
        "sender_location": "Istanbul",
        "receiver_location": "London",
    },
    139: {
        "date": "1859-02-09",
        "sender": "Ottoman Foreign Minister Keçecizade Fuad Pasha",
        "receiver": "Ottoman Consulate of New York, Hosford Smith",
        "sender_location": "Istanbul",
        "receiver_location": "New York",
    },
    140: {
        "date": "1859-03-15",
        "sender": "Ottoman Consulate of New York, Hosford Smith",
        "receiver": "Ottoman Foreign Minister Keçecizade Fuad Pasha",
        "sender_location": "New York",
        "receiver_location": "Istanbul",
    },
}


def _comm_id_to_int(cid: str) -> int:
    return int(cid.split("_", 1)[1])


def patch_communications() -> None:
    """Apply CORRECTIONS to data/communications.json.

    We rely on the existing date-based mapping in build_documents.py to know
    which comm entry currently lives at each Dosya. For Dosya numbers whose
    target date already has a free comm we update in place; otherwise we
    append a new comm_<n> entry. We keep entries chronologically ordered to
    match the PURE workbook convention.
    """
    docs = json.loads((ROOT / "data" / "diplomatic_documents.json").read_text(encoding="utf-8"))
    comms = json.loads(COMMS_PATH.read_text(encoding="utf-8"))

    dosya_to_comm: dict[int, str] = {}
    for d in docs:
        if d.get("communication_id") and d["id"].startswith("diplomatic_"):
            n = int(d["id"].split("_", 1)[1])
            dosya_to_comm[n] = d["communication_id"]

    used_comm_ids = set()
    by_id = {c["id"]: c for c in comms}

    def _next_comm_id() -> str:
        n = 0
        while True:
            cid = f"comm_{n}"
            if cid not in by_id and cid not in used_comm_ids:
                return cid
            n += 1

    for dosya_num, fix in CORRECTIONS.items():
        target_date = fix["date"]
        existing_id = dosya_to_comm.get(dosya_num)
        entry: dict | None = None
        if existing_id and existing_id in by_id:
            entry = by_id[existing_id]
        else:
            # Append a new comm_<n> entry; pick the lowest unused index.
            new_id = _next_comm_id()
            entry = {
                "id": new_id,
                "date": target_date,
                "sender": "",
                "receiver": "",
                "sender_location": "",
                "receiver_location": "",
                "type": "letter",
            }
            comms.append(entry)
            by_id[new_id] = entry
            used_comm_ids.add(new_id)

        entry["date"] = target_date
        entry["sender"] = fix["sender"]
        entry["receiver"] = fix["receiver"]
        entry["sender_location"] = fix["sender_location"]
        entry["receiver_location"] = fix["receiver_location"]
        entry["type"] = fix.get("type", entry.get("type", "letter"))

    # Keep entries chronologically sorted; preserve stable id when dates tie.
    comms.sort(key=lambda c: (c.get("date") or "", _comm_id_to_int(c["id"])))

    COMMS_PATH.write_text(
        json.dumps(comms, ensure_ascii=False, indent=2) + "\n",
        encoding="utf-8",
    )
    print(f"[comms] wrote {COMMS_PATH} ({len(comms)} entries)")


def patch_excel() -> None:
    """Mirror the corrections into ``Pace of Communication - PURE.xlsx``.

    The sheet is keyed by date + sender + receiver, so we re-derive its
    content from the freshly-patched communications.json. This keeps the
    spreadsheet and JSON in lock-step.
    """
    if not EXCEL_PATH.exists():
        print(f"[excel] {EXCEL_PATH} missing; skipping spreadsheet sync")
        return

    comms = json.loads(COMMS_PATH.read_text(encoding="utf-8"))
    wb = openpyxl.load_workbook(EXCEL_PATH)
    ws = wb.active
    if ws is None:
        print(f"[excel] {EXCEL_PATH} has no active sheet; skipping")
        return

    header = [c.value for c in ws[1]]
    # Clear data rows
    if ws.max_row > 1:
        ws.delete_rows(2, ws.max_row - 1)

    def _iso_to_excel(iso: str) -> str:
        y, m, d = iso.split("-")
        return f"{d}.{m}.{y}"

    for c in comms:
        ws.append([
            _iso_to_excel(c["date"]) if c.get("date") else "",
            c.get("sender") or "",
            c.get("sender_location") or "",
            c.get("receiver") or "",
            c.get("receiver_location") or "",
        ])

    wb.save(EXCEL_PATH)
    print(f"[excel] wrote {EXCEL_PATH} ({len(comms)} rows)")


def main() -> None:
    patch_communications()
    patch_excel()


if __name__ == "__main__":
    main()
