"""
Generates networkx-map.html from Network map.xlsx
Run: python3 scripts/build_network_map.py
"""

import openpyxl
from pyvis.network import Network
from pathlib import Path

EXCEL = Path("/Users/odulsuzkisafilm/Downloads/Network map.xlsx")
OUTPUT = Path(__file__).parent.parent / "networkx-map.html"

# ── Colours ──────────────────────────────────────────────────────────────────
COLORS = {
    "Criminal": {
        "background": "#c0392b",   # deep red
        "border":     "#7b241c",
        "highlight":  {"background": "#e8dcc5", "border": "#7b241c"},
        "hover":      {"background": "#e8dcc5", "border": "#7b241c"},
    },
    "City": {
        "background": "#1a2a40",   # dark navy (matches site primary)
        "border":     "#0d1520",
        "highlight":  {"background": "#a67b5b", "border": "#0d1520"},
        "hover":      {"background": "#a67b5b", "border": "#0d1520"},
    },
}

EDGE_COLOR_ASSOCIATES = "rgba(192, 57, 43, 0.35)"   # translucent red
EDGE_COLOR_LOCATED    = "rgba(26,  42, 64, 0.45)"   # translucent navy

# ── Load Excel ────────────────────────────────────────────────────────────────
wb = openpyxl.load_workbook(EXCEL)

nodes_ws = wb["Nodes"]
edges_ws = wb["Edges"]

nodes = {}  # id → {label, type}
for row in nodes_ws.iter_rows(min_row=2, values_only=True):
    nid, label, ntype = row
    if nid:
        nodes[str(nid).strip()] = {"label": str(label).strip(), "type": str(ntype).strip()}

edges = []
for row in edges_ws.iter_rows(min_row=2, values_only=True):
    src, tgt, etype = row
    if src and tgt:
        edges.append((str(src).strip(), str(tgt).strip(), str(etype).strip()))

# ── Build degree map for node sizing ─────────────────────────────────────────
degree = {nid: 0 for nid in nodes}
for src, tgt, _ in edges:
    if src != tgt:          # skip self-loops
        degree[src] = degree.get(src, 0) + 1
        degree[tgt] = degree.get(tgt, 0) + 1

# ── Build pyvis network ───────────────────────────────────────────────────────
net = Network(
    height="700px",
    width="100%",
    bgcolor="#f8f5ef",
    font_color="#2c3440",
    directed=False,
    notebook=False,
)

net.set_options("""
{
  "nodes": {
    "font": { "size": 13, "face": "Montserrat, sans-serif" },
    "borderWidth": 2,
    "borderWidthSelected": 3
  },
  "edges": {
    "smooth": { "type": "continuous" },
    "width": 1.2
  },
  "physics": {
    "forceAtlas2Based": {
      "gravitationalConstant": -80,
      "centralGravity": 0.01,
      "springLength": 120,
      "springConstant": 0.06,
      "damping": 0.5,
      "avoidOverlap": 0.4
    },
    "solver": "forceAtlas2Based",
    "stabilization": { "iterations": 200 }
  },
  "interaction": {
    "hover": true,
    "tooltipDelay": 100,
    "hideEdgesOnDrag": false,
    "navigationButtons": true,
    "keyboard": true
  }
}
""")

for nid, info in nodes.items():
    ntype = info["type"]
    color = COLORS.get(ntype, COLORS["Criminal"])
    deg   = degree.get(nid, 0)

    if ntype == "City":
        size  = 28
        shape = "diamond"
        font_size = 15
    else:
        # scale Criminal nodes between 12 and 32 by degree
        size  = max(12, min(32, 10 + deg * 1.5))
        shape = "dot"
        font_size = 12

    # Plain text title — HTML will be injected via JS post-processing
    title = f"{info['label']} | {ntype} | Connections: {deg}"

    net.add_node(
        nid,
        label=info["label"],
        title=title,
        color=color,
        size=size,
        shape=shape,
        font={"size": font_size, "color": "#2c3440"},
    )

for src, tgt, etype in edges:
    if src == tgt:
        continue  # skip self-loops
    if etype == "located_in":
        color = EDGE_COLOR_LOCATED
        dashes = True
        width  = 1.0
    else:
        color = EDGE_COLOR_ASSOCIATES
        dashes = False
        width  = 1.5

    net.add_edge(src, tgt, color=color, dashes=dashes, width=width, title=etype)

# ── Write output ──────────────────────────────────────────────────────────────
net.save_graph(str(OUTPUT))

# ── Post-process: inject selectNode / deselectNode dimming + HTML tooltips ────
INTERACTION_JS = r"""
                  // ── Search controls ───────────────────────────────────────────
                  var _searchInput = document.getElementById('network-node-search');
                  var _searchBtn = document.getElementById('network-search-btn');
                  var _clearBtn = document.getElementById('network-clear-btn');
                  var _searchList = document.getElementById('network-node-list');

                  function _populateSearchList() {
                      if (!_searchList) return;
                      _searchList.innerHTML = '';
                      nodes.get().forEach(function(n) {
                          var opt = document.createElement('option');
                          opt.value = n.label;
                          _searchList.appendChild(opt);
                      });
                  }

                  function _findNodeByQuery(query) {
                      var q = (query || '').trim().toLowerCase();
                      if (!q) return null;
                      var all = nodes.get();
                      var exact = all.find(function(n) { return (n.label || '').toLowerCase() === q; });
                      if (exact) return exact;
                      return all.find(function(n) { return (n.label || '').toLowerCase().indexOf(q) !== -1; }) || null;
                  }

                  function _focusNodeByQuery(query) {
                      var hit = _findNodeByQuery(query);
                      if (!hit) return false;
                      network.selectNodes([hit.id]);
                      network.focus(hit.id, { scale: 1.2, animation: { duration: 450, easingFunction: 'easeInOutQuad' } });
                      return true;
                  }

                  function _clearSearchAndSelection() {
                      if (_searchInput) _searchInput.value = '';
                      network.unselectAll();
                  }

                  if (_searchBtn) {
                      _searchBtn.addEventListener('click', function() {
                          if (!_focusNodeByQuery(_searchInput ? _searchInput.value : '')) {
                              if (_searchInput) _searchInput.setCustomValidity('Node not found');
                              if (_searchInput) _searchInput.reportValidity();
                              if (_searchInput) _searchInput.setCustomValidity('');
                          }
                      });
                  }
                  if (_searchInput) {
                      _searchInput.addEventListener('keydown', function(evt) {
                          if (evt.key === 'Enter') {
                              evt.preventDefault();
                              if (!_focusNodeByQuery(_searchInput.value)) {
                                  _searchInput.setCustomValidity('Node not found');
                                  _searchInput.reportValidity();
                                  _searchInput.setCustomValidity('');
                              }
                          }
                      });
                  }
                  if (_clearBtn) {
                      _clearBtn.addEventListener('click', _clearSearchAndSelection);
                  }

                  // ── Node selection: dim non-neighbours ──────────────────────
                  var _origColors = {};
                  var _origEdgeColors = {};

                  network.once('stabilized', function() {
                      _populateSearchList();
                      nodes.get().forEach(function(n) {
                          _origColors[n.id] = JSON.parse(JSON.stringify(n.color));
                      });
                      edges.get().forEach(function(e) {
                          _origEdgeColors[e.id] = e.color;
                      });
                  });

                  network.on('selectNode', function(params) {
                      var sel = params.nodes[0];
                      var neighbors = new Set(network.getConnectedNodes(sel));
                      neighbors.add(sel);

                      var nodeUpdates = [];
                      nodes.get().forEach(function(n) {
                          var dimmed = !neighbors.has(n.id);
                          var c = JSON.parse(JSON.stringify(_origColors[n.id] || n.color));
                          if (dimmed) {
                              c.background = 'rgba(200,200,200,0.25)';
                              c.border     = 'rgba(180,180,180,0.25)';
                          }
                          nodeUpdates.push({ id: n.id, color: c, opacity: dimmed ? 0.2 : 1,
                              font: { color: dimmed ? 'rgba(180,180,180,0.3)' : '#2c3440' } });
                      });
                      nodes.update(nodeUpdates);

                      var edgeUpdates = [];
                      edges.get().forEach(function(e) {
                          var connected = neighbors.has(e.from) && neighbors.has(e.to);
                          edgeUpdates.push({ id: e.id,
                              color: connected ? (_origEdgeColors[e.id] || e.color) : 'rgba(200,200,200,0.1)' });
                      });
                      edges.update(edgeUpdates);
                  });

                  network.on('deselectNode', function() {
                      var nodeUpdates = [];
                      nodes.get().forEach(function(n) {
                          nodeUpdates.push({ id: n.id,
                              color: _origColors[n.id] || n.color,
                              opacity: 1,
                              font: { color: '#2c3440' } });
                      });
                      nodes.update(nodeUpdates);

                      var edgeUpdates = [];
                      edges.get().forEach(function(e) {
                          edgeUpdates.push({ id: e.id, color: _origEdgeColors[e.id] || e.color });
                      });
                      edges.update(edgeUpdates);
                  });

                  // ── HTML tooltips via DOM elements ───────────────────────────
                  network.on('showPopup', function(nodeId) {
                      var n = nodes.get(nodeId);
                      if (!n) return;
                      setTimeout(function() {
                          var tip = document.querySelector('.vis-tooltip');
                          if (!tip || !n.title) return;
                          var parts = n.title.split(' | ');
                          tip.innerHTML =
                              '<div style="font-family:Montserrat,sans-serif;padding:4px 2px;">' +
                              '<strong style="font-size:13px;">' + (parts[0] || '') + '</strong>' +
                              (parts[1] ? '<br><span style="color:#888;font-size:11px;">' + parts[1] + '</span>' : '') +
                              (parts[2] ? '<br><span style="font-size:11px;">' + parts[2] + '</span>' : '') +
                              '</div>';
                      }, 0);
                  });
"""

html = OUTPUT.read_text(encoding="utf-8")
html = html.replace(
    '<div id="mynetwork" class="card-body"></div>',
    """<div style="padding: 10px 12px 0 12px; display:flex; gap:8px; align-items:center; flex-wrap:wrap;">
                <input id="network-node-search" list="network-node-list" type="text" placeholder="Search node..." style="min-width:220px; padding:6px 10px; border:1px solid #c7c0b5; border-radius:6px; font-family:Montserrat,sans-serif; font-size:13px;">
                <datalist id="network-node-list"></datalist>
                <button id="network-search-btn" type="button" style="padding:6px 10px; border:1px solid #1a2a40; background:#1a2a40; color:#fff; border-radius:6px; font-size:12px; font-family:Montserrat,sans-serif;">Find</button>
                <button id="network-clear-btn" type="button" style="padding:6px 10px; border:1px solid #c7c0b5; background:#fff; color:#333; border-radius:6px; font-size:12px; font-family:Montserrat,sans-serif;">Clear</button>
              </div>
              <div id="mynetwork" class="card-body"></div>"""
)
html = html.replace("                  return network;", INTERACTION_JS + "\n                  return network;")
OUTPUT.write_text(html, encoding="utf-8")

print(f"✓  Written → {OUTPUT}")
print(f"   {len(nodes)} nodes, {len(edges)} edges")
